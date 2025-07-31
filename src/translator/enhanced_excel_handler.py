"""增强的Excel文件处理工具 - 支持合并单元格识别和重建."""

import os
import pandas as pd
from typing import Dict, List, Any
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import tempfile
from config.logging_config import get_logger

logger = get_logger(__name__)


class EnhancedExcelHandler:
    """增强的Excel处理类，支持合并单元格识别和重建."""

    def __init__(self):
        return

    def read_excel_with_merge_info(self, file_path: str) -> Dict[str, Any]:
        """
        读取Excel文件并保留合并单元格信息和样式信息.

        Args:
            file_path: Excel文件路径

        Returns:
            Dictionary with 'data' (原始数据), 'merge_info' (合并单元格信息) and 'style_info' (样式信息)
        """
        try:
            excel_data = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
            merge_info = self._get_merge_info(file_path)
            style_info = self._get_style_info(file_path)
            return {
                "data": excel_data,
                "merge_info": merge_info,
                "style_info": style_info,
            }
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")

    def _get_merge_info(self, file_path: str) -> Dict[str, List]:
        """获取所有sheet的合并单元格信息."""
        workbook = load_workbook(file_path)
        merge_info = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            merge_info[sheet_name] = []
            for merged_range in worksheet.merged_cells.ranges:
                merge_info[sheet_name].append(
                    {
                        "range": str(merged_range),
                        "min_row": merged_range.min_row,
                        "max_row": merged_range.max_row,
                        "min_col": merged_range.min_col,
                        "max_col": merged_range.max_col,
                    }
                )
        return merge_info

    def _get_style_info(self, file_path: str) -> Dict[str, Any]:
        """获取样式信息（用于保持格式一致）."""
        workbook = load_workbook(file_path)
        style_info = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            style_info[sheet_name] = {
                "column_widths": {},
                "row_heights": {},
                "cell_alignment": {},
            }
            for col_letter, col_dimension in worksheet.column_dimensions.items():
                style_info[sheet_name]["column_widths"][col_letter] = (
                    col_dimension.width
                )
            for row_num, row_dimension in worksheet.row_dimensions.items():
                style_info[sheet_name]["row_heights"][row_num] = row_dimension.height
            for row in worksheet.iter_rows():
                for cell in row:
                    if (
                        hasattr(cell, "is_merged")
                        and cell.is_merged
                        and (
                            cell.coordinate
                            != worksheet.cell(cell.row, cell.column).coordinate
                        )
                    ):
                        continue
                    if cell.alignment is not None:
                        if (
                            cell.alignment.horizontal != "general"
                            or cell.alignment.vertical != "bottom"
                            or cell.alignment.text_rotation != 0
                            or (cell.alignment.wrap_text is not False)
                            or (cell.alignment.shrink_to_fit is not False)
                            or (cell.alignment.indent != 0)
                        ):
                            cell_key = cell.coordinate
                            style_info[sheet_name]["cell_alignment"][cell_key] = {
                                "horizontal": cell.alignment.horizontal,
                                "vertical": cell.alignment.vertical,
                                "text_rotation": cell.alignment.text_rotation,
                                "wrap_text": cell.alignment.wrap_text,
                                "shrink_to_fit": cell.alignment.shrink_to_fit,
                                "indent": cell.alignment.indent,
                            }
        return style_info

    def write_excel_with_merge_info(
        self,
        original_file_path: str,
        translated_data: Dict[str, pd.DataFrame],
        output_path: str,
        merge_info: Dict[str, List],
        style_info: Dict[str, Any] = None,
    ) -> str:
        """
        写入Excel文件并重建合并单元格.

        Args:
            original_file_path: 原始文件路径（用于复制格式）
            translated_data: 翻译后的数据
            output_path: 输出文件路径
            merge_info: 合并单元格信息
            style_info: 样式信息（可选）

        Returns:
            输出文件路径
        """
        try:
            template_wb = load_workbook(original_file_path)
            new_wb = Workbook()
            new_wb.remove(new_wb.active)
            for sheet_name, df in translated_data.items():
                if sheet_name in new_wb.sheetnames:
                    new_ws = new_wb[sheet_name]
                else:
                    new_ws = new_wb.create_sheet(sheet_name)
                if sheet_name in template_wb.sheetnames:
                    template_ws = template_wb[sheet_name]
                    self._copy_worksheet_format(template_ws, new_ws)
                self._write_data_to_worksheet(new_ws, df)
                if sheet_name in merge_info:
                    self._recreate_merged_cells(new_ws, merge_info[sheet_name])
                if style_info and sheet_name in style_info:
                    self._apply_style_info(new_ws, style_info[sheet_name])
            new_wb.save(output_path)
            return output_path
        except Exception as e:
            raise Exception(f"写入Excel文件失败: {str(e)}")

    def _copy_worksheet_format(self, source_ws: Worksheet, target_ws: Worksheet):
        """复制工作表格式."""
        for col_letter, col_dimension in source_ws.column_dimensions.items():
            if col_dimension.width is not None:
                target_ws.column_dimensions[col_letter].width = col_dimension.width
        for row_num, row_dimension in source_ws.row_dimensions.items():
            if row_dimension.height is not None:
                target_ws.row_dimensions[row_num].height = row_dimension.height
        from openpyxl.styles import Alignment

        for row in source_ws.iter_rows():
            for cell in row:
                if (
                    hasattr(cell, "is_merged")
                    and cell.is_merged
                    and (
                        cell.coordinate
                        != source_ws.cell(cell.row, cell.column).coordinate
                    )
                ):
                    continue
                if cell.alignment is not None:
                    if (
                        cell.alignment.horizontal != "general"
                        or cell.alignment.vertical != "bottom"
                        or cell.alignment.text_rotation != 0
                        or (cell.alignment.wrap_text is not False)
                        or (cell.alignment.shrink_to_fit is not False)
                        or (cell.alignment.indent != 0)
                    ):
                        target_cell = target_ws[cell.coordinate]
                        target_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent,
                        )

    def _write_data_to_worksheet(self, worksheet: Worksheet, df: pd.DataFrame):
        """将DataFrame数据写入工作表."""
        for col_idx, column_name in enumerate(df.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=column_name)
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

    def _recreate_merged_cells(self, worksheet: Worksheet, merge_info: List):
        """重建合并单元格."""
        for merge_data in merge_info:
            try:
                merge_range = f"{get_column_letter(merge_data['min_col'])}{merge_data['min_row']}:{get_column_letter(merge_data['max_col'])}{merge_data['max_row']}"
                worksheet.merge_cells(merge_range)
            except Exception as e:
                logger.warning(f"重建合并单元格失败: {merge_range}, 错误: {str(e)}")

    def _apply_style_info(self, worksheet: Worksheet, style_info: Dict[str, Any]):
        """应用样式信息."""
        if "column_widths" in style_info:
            for col_letter, width in style_info["column_widths"].items():
                if width is not None:
                    worksheet.column_dimensions[col_letter].width = width
        if "row_heights" in style_info:
            for row_num, height in style_info["row_heights"].items():
                if height is not None:
                    worksheet.row_dimensions[row_num].height = height
        if "cell_alignment" in style_info:
            from openpyxl.styles import Alignment

            for cell_key, alignment_info in style_info["cell_alignment"].items():
                cell = worksheet[cell_key]
                cell.alignment = Alignment(
                    horizontal=alignment_info.get("horizontal", "general"),
                    vertical=alignment_info.get("vertical", "bottom"),
                    text_rotation=alignment_info.get("text_rotation", 0),
                    wrap_text=alignment_info.get("wrap_text", False),
                    shrink_to_fit=alignment_info.get("shrink_to_fit", False),
                    indent=alignment_info.get("indent", 0),
                )

    def read_excel(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Read Excel file with multiple sheets.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary with sheet names as keys and DataFrames as values
        """
        try:
            excel_data = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
            processed_data = {}
            for sheet_name, df in excel_data.items():
                df = self._standardize_columns(df)
                processed_data[sheet_name] = df
            return processed_data
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")

    def write_excel(self, data: Dict[str, pd.DataFrame], output_path: str) -> str:
        """
        Write translated data to Excel file.

        Args:
            data: Dictionary with sheet names and translated DataFrames
            output_path: Output file path

        Returns:
            Path to the created file
        """
        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for sheet_name, df in data.items():
                    clean_df = df.fillna("")
                    clean_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = None
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                            if column_letter is None:
                                column_letter = cell.column_letter
                        if column_letter:
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[
                                column_letter
                            ].width = adjusted_width
                return output_path
        except Exception as e:
            raise Exception(f"写入Excel文件失败: {str(e)}")

    def get_sheet_info(self, file_path: str) -> Dict[str, Any]:
        """Get information about Excel sheets."""
        try:
            workbook = load_workbook(file_path)
            info = {"sheet_names": list(workbook.sheetnames), "sheet_details": {}}
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                info["sheet_details"][sheet_name] = {
                    "rows": worksheet.max_row,
                    "columns": worksheet.max_column,
                    "headers": [
                        worksheet.cell(row=1, column=col).value
                        for col in range(1, min(5, worksheet.max_column + 1))
                    ],
                }
            return info
        except Exception as e:
            raise Exception(f"获取sheet信息失败: {str(e)}")

    def create_temp_copy(self, file_path: str) -> str:
        """Create a temporary copy of the file for processing."""
        try:
            temp_dir = tempfile.mkdtemp()
            temp_path = os.path.join(temp_dir, os.path.basename(file_path))
            import shutil

            shutil.copy2(file_path, temp_path)
            return temp_path
        except Exception as e:
            raise Exception(f"创建临时文件失败: {str(e)}")

    def extract_text_for_translation(
        self, data: Dict[str, pd.DataFrame]
    ) -> Dict[str, List[str]]:
        """
        Extract all text content for translation from all columns.

        Args:
            data: Dictionary with sheet names and DataFrames

        Returns:
            Dictionary with sheet names and lists of text to translate
        """
        text_to_translate = {}
        for sheet_name, df in data.items():
            texts = []
            for column in df.columns:
                texts.extend(df[column].dropna().astype(str).tolist())
            text_to_translate[sheet_name] = texts
        return text_to_translate

    def apply_translations(
        self, data: Dict[str, pd.DataFrame], translations: Dict[str, Dict[str, str]]
    ) -> Dict[str, pd.DataFrame]:
        """
        Apply translations back to all DataFrame columns.

        Args:
            data: Original data
            translations: Dictionary with translations for each sheet

        Returns:
            Updated DataFrames with translated content
        """
        translated_data = {}
        for sheet_name, df in data.items():
            translated_df = df.copy()
            if sheet_name in translations:
                sheet_translations = translations[sheet_name]
                for column in translated_df.columns:

                    def translate_value(x):
                        if pd.isna(x) or x is None or str(x).strip() == "":
                            return x
                        return sheet_translations.get(str(x), str(x))

                    translated_df[column] = translated_df[column].apply(translate_value)
            translated_data[sheet_name] = translated_df
        return translated_data

    def _standardize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        标准化列名，处理空列名或"Unnamed"列名的问题.

        Args:
            df: 输入的DataFrame

        Returns:
            处理后的DataFrame
        """
        # 复制DataFrame以避免修改原始数据
        df_copy = df.copy()

        # 处理列名中的"Unnamed"问题
        new_columns = []
        for col in df_copy.columns:
            # 如果列名是"Unnamed: 0"或类似的未命名列，保持为空或设置为空字符串
            if isinstance(col, str) and col.startswith("Unnamed:"):
                new_columns.append("")
            else:
                new_columns.append(col)

        df_copy.columns = new_columns
        return df_copy
