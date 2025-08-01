from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment
import json


def html_to_excel_with_format(html_file, excel_file, meta_file=None):
    """
    HTML转Excel，恢复原始格式，支持多sheet
    """
    # 读取HTML文件
    with open(html_file, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "html.parser")

    # 读取格式元数据
    all_format_info = None
    if meta_file:
        with open(meta_file, "r", encoding="utf-8") as f:
            all_format_info = json.load(f)

    # 创建工作簿
    wb = Workbook()
    # 删除默认工作表
    wb.remove(wb.active)

    # 获取所有工作表容器
    sheet_containers = soup.find_all("div", class_="sheet-container")
    
    # 如果没有找到工作表容器，尝试兼容旧格式
    if not sheet_containers:
        # 兼容旧格式 - 创建一个默认工作表
        ws = wb.create_sheet("Sheet1")
        table = soup.find("table")
        if not table:
            raise ValueError("HTML中未找到表格")
        
        format_info = all_format_info.get("Sheet1", None) if all_format_info else None
        _process_table_data(ws, table, format_info)
    else:
        # 处理每个工作表
        for container in sheet_containers:
            sheet_name = container.get("data-sheet-name", f"Sheet{len(wb.sheetnames) + 1}")
            # 创建工作表
            ws = wb.create_sheet(sheet_name)
            
            # 获取表格数据
            table = container.find("table")
            if not table:
                raise ValueError(f"工作表 {sheet_name} 中未找到表格")
            
            # 获取当前工作表的格式信息
            format_info = all_format_info.get(sheet_name, None) if all_format_info else None
            
            # 处理表格数据
            _process_table_data(ws, table, format_info)

    # 保存Excel文件
    wb.save(excel_file)


def _process_table_data(ws, table, format_info=None):
    """
    处理单个工作表的表格数据
    """
    rows = table.find_all("tr")

    # 解析HTML表格数据
    excel_data = []
    for i, row in enumerate(rows):
        cells = row.find_all(["td", "th"])
        row_data = []
        for j, cell in enumerate(cells):
            cell_data = {
                "value": cell.get_text().strip(),
                "colspan": int(cell.get("colspan", 1)),
                "rowspan": int(cell.get("rowspan", 1)),
                "style": cell.get("style", ""),
            }
            row_data.append(cell_data)
        excel_data.append(row_data)

    # 写入Excel数据
    merged_cells_info = []
    for i, row_data in enumerate(excel_data):
        col_offset = 0
        for j, cell_data in enumerate(row_data):
            # 计算实际列位置（考虑合并单元格的影响）
            actual_col = j + 1 + col_offset

            # 写入单元格值
            # 检查是否是合并单元格的非左上角单元格，如果是则跳过
            is_merged_cell = False
            if format_info and "merged_cells" in format_info:
                for merged_range in format_info["merged_cells"]:
                    if (
                        merged_range["min_row"] <= i + 1 <= merged_range["max_row"]
                        and merged_range["min_col"]
                        <= actual_col
                        <= merged_range["max_col"]
                    ):
                        # 如果是合并单元格的非左上角单元格，则跳过
                        if not (
                            i + 1 == merged_range["min_row"]
                            and actual_col == merged_range["min_col"]
                        ):
                            is_merged_cell = True
                            break
            elif merged_cells_info:
                # 如果没有元数据，从HTML中恢复合并单元格信息
                for start_row, start_col, end_row, end_col in merged_cells_info:
                    if (
                        start_row <= i + 1 <= end_row
                        and start_col <= actual_col <= end_col
                    ):
                        # 如果是合并单元格的非左上角单元格，则跳过
                        if not (i + 1 == start_row and actual_col == start_col):
                            is_merged_cell = True
                            break

            if not is_merged_cell:
                # 处理pandas生成的"Unnamed:"列名，将其替换为空字符串
                cell_value = cell_data["value"]
                if cell_value.startswith("Unnamed:"):
                    cell_value = ""
                ws.cell(row=i + 1, column=actual_col, value=cell_value)

            # 处理合并单元格
            if cell_data["colspan"] > 1 or cell_data["rowspan"] > 1:
                end_col = actual_col + cell_data["colspan"] - 1
                end_row = i + 1 + cell_data["rowspan"] - 1
                ws.merge_cells(
                    start_row=i + 1,
                    start_column=actual_col,
                    end_row=end_row,
                    end_column=end_col,
                )
                merged_cells_info.append((i + 1, actual_col, end_row, end_col))

            # 更新列偏移
            col_offset += cell_data["colspan"] - 1

            # 应用样式（如果有元数据）
            if format_info and f"{i + 1}-{actual_col}" in format_info["cell_styles"]:
                cell_style = format_info["cell_styles"][f"{i + 1}-{actual_col}"]
                # 检查是否是合并单元格的非左上角单元格，如果是则跳过样式应用
                is_merged_cell = False
                if "merged_cells" in format_info:
                    for merged_range in format_info["merged_cells"]:
                        if (
                            merged_range["min_row"] <= i + 1 <= merged_range["max_row"]
                            and merged_range["min_col"]
                            <= actual_col
                            <= merged_range["max_col"]
                        ):
                            # 如果是合并单元格的非左上角单元格，则跳过
                            if not (
                                i + 1 == merged_range["min_row"]
                                and actual_col == merged_range["min_col"]
                            ):
                                is_merged_cell = True
                                break

                if not is_merged_cell:
                    cell = ws.cell(row=i + 1, column=actual_col)

                # 应用字体样式（已移除）

                # 应用背景色（已移除）

                # 应用对齐方式
                if "alignment" in cell_style:
                    alignment = Alignment(
                        horizontal=cell_style["alignment"]["horizontal"],
                        vertical=cell_style["alignment"]["vertical"],
                    )
                    cell.alignment = alignment

    # 恢复合并单元格（如果没有元数据，则从HTML中恢复）
    if not format_info and merged_cells_info:
        for start_row, start_col, end_row, end_col in merged_cells_info:
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col,
            )

    # 恢复行高和列宽
    if format_info:
        # 恢复行高
        for row_num, height in format_info["row_heights"].items():
            if height:
                ws.row_dimensions[int(row_num)].height = float(height)

        # 恢复列宽
        for col_letter, width in format_info["col_widths"].items():
            if width:
                ws.column_dimensions[col_letter].width = float(width)
