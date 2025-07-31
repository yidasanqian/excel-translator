"""Excel file handling utilities."""

import os
import pandas as pd
from typing import Dict, List, Any
from openpyxl import load_workbook
import tempfile


class ExcelHandler:
    """Handles Excel file reading and writing operations."""

    def __init__(self):
        return

    def read_excel(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Read Excel file with multiple sheets.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary with sheet names as keys and DataFrames as values
        """
        try:
            return pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")

    def write_excel(self, data: Dict[str, pd.DataFrame], output_path: str) -> str:
        """
        Write translated data to Excel file.

        Args:
            data: Dictionary with sheet names and translated DataFrames
            output_path: Output file path

        Returns:
            Path to the created file
        """
        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for sheet_name, df in data.items():
                    clean_df = df.fillna("")
                    clean_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = None
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                            if column_letter is None:
                                column_letter = cell.column_letter
                        if column_letter:
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[
                                column_letter
                            ].width = adjusted_width
                return output_path
        except Exception as e:
            raise Exception(f"写入Excel文件失败: {str(e)}")

    def get_sheet_info(self, file_path: str) -> Dict[str, Any]:
        """Get information about Excel sheets."""
        try:
            workbook = load_workbook(file_path)
            info = {"sheet_names": list(workbook.sheetnames), "sheet_details": {}}
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                info["sheet_details"][sheet_name] = {
                    "rows": worksheet.max_row,
                    "columns": worksheet.max_column,
                    "headers": [
                        worksheet.cell(row=1, column=col).value
                        for col in range(1, min(5, worksheet.max_column + 1))
                    ],
                }
            return info
        except Exception as e:
            raise Exception(f"获取sheet信息失败: {str(e)}")

    def create_temp_copy(self, file_path: str) -> str:
        """Create a temporary copy of the file for processing."""
        try:
            temp_dir = tempfile.mkdtemp()
            temp_path = os.path.join(temp_dir, os.path.basename(file_path))
            import shutil

            shutil.copy2(file_path, temp_path)
            return temp_path
        except Exception as e:
            raise Exception(f"创建临时文件失败: {str(e)}")

    def extract_text_for_translation(
        self, data: Dict[str, pd.DataFrame]
    ) -> Dict[str, List[str]]:
        """
        Extract all text content for translation from all columns.

        Args:
            data: Dictionary with sheet names and DataFrames

        Returns:
            Dictionary with sheet names and lists of text to translate
        """
        text_to_translate = {}
        for sheet_name, df in data.items():
            texts = []
            for column in df.columns:
                texts.extend(df[column].dropna().astype(str).tolist())
            text_to_translate[sheet_name] = texts
        return text_to_translate

    def apply_translations(
        self, data: Dict[str, pd.DataFrame], translations: Dict[str, Dict[str, str]]
    ) -> Dict[str, pd.DataFrame]:
        """
        Apply translations back to all DataFrame columns.

        Args:
            data: Original data
            translations: Dictionary with translations for each sheet

        Returns:
            Updated DataFrames with translated content
        """
        translated_data = {}
        for sheet_name, df in data.items():
            translated_df = df.copy()
            if sheet_name in translations:
                sheet_translations = translations[sheet_name]
                for column in translated_df.columns:

                    def translate_value(x):
                        if pd.isna(x) or x is None or str(x).strip() == "":
                            return x
                        return sheet_translations.get(str(x), str(x))

                    translated_df[column] = translated_df[column].apply(translate_value)
            translated_data[sheet_name] = translated_df
        return translated_data
