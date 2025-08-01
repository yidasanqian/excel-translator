from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def excel_to_html_with_format(excel_file, html_file, meta_file=None):
    """
    Excel转HTML，保持格式并保存元数据，支持多sheet
    """
    wb = load_workbook(excel_file)

    # 存储所有工作表的格式信息
    all_format_info = {}

    # 存储所有工作表的HTML内容
    all_html_content = []

    # 处理每个工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 为当前工作表存储格式信息
        format_info = {
            "merged_cells": [],
            "cell_styles": {},
            "row_heights": {},
            "col_widths": {},
        }

        # 获取合并单元格信息
        for merged_range in ws.merged_cells.ranges:
            format_info["merged_cells"].append(
                {
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col,
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                }
            )

        # 获取行高和列宽
        for row in range(1, ws.max_row + 1):
            format_info["row_heights"][row] = ws.row_dimensions[row].height

        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            format_info["col_widths"][col_letter] = ws.column_dimensions[
                col_letter
            ].width

        html_content = f"""
    <div class="sheet-container" data-sheet-name="{sheet_name}">
        <h2 class="sheet-title">{sheet_name}</h2>
        <table class="excel-table" data-sheet-name="{sheet_name}">
"""

        # 生成HTML表格
        for row_num in range(1, ws.max_row + 1):
            html_content += "            <tr>\n"
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)

                # 记录单元格样式（只保留对齐样式）
                cell_style = {
                    "alignment": {
                        "horizontal": str(cell.alignment.horizontal)
                        if cell.alignment.horizontal
                        else "left",
                        "vertical": str(cell.alignment.vertical)
                        if cell.alignment.vertical
                        else "bottom",
                    },
                }
                format_info["cell_styles"][f"{row_num}-{col_num}"] = cell_style

                # 处理合并单元格
                colspan = ""
                rowspan = ""
                skip_cell = False

                for merged_range in format_info["merged_cells"]:
                    if (
                        row_num == merged_range["min_row"]
                        and col_num == merged_range["min_col"]
                    ):
                        if merged_range["max_col"] > merged_range["min_col"]:
                            colspan = f' colspan="{merged_range["max_col"] - merged_range["min_col"] + 1}"'
                        if merged_range["max_row"] > merged_range["min_row"]:
                            rowspan = f' rowspan="{merged_range["max_row"] - merged_range["min_row"] + 1}"'
                    elif (
                        merged_range["min_row"] <= row_num <= merged_range["max_row"]
                        and merged_range["min_col"]
                        <= col_num
                        <= merged_range["max_col"]
                        and not (
                            row_num == merged_range["min_row"]
                            and col_num == merged_range["min_col"]
                        )
                    ):
                        skip_cell = True
                        break

                if skip_cell:
                    continue

                # 构建样式（只保留对齐样式）
                style_parts = []
                if cell_style["alignment"]["horizontal"]:
                    style_parts.append(
                        f"text-align: {cell_style['alignment']['horizontal']}"
                    )
                if cell_style["alignment"]["vertical"]:
                    style_parts.append(
                        f"vertical-align: {cell_style['alignment']['vertical']}"
                    )

                style_attr = f' style="{"; ".join(style_parts)}"' if style_parts else ""

                cell_value = cell.value if cell.value is not None else ""
                html_content += f"                <td{colspan}{rowspan}{style_attr}>{cell_value}</td>\n"

            html_content += "            </tr>\n"

        html_content += """        </table>
    </div>"""

        # 保存当前工作表的信息
        all_format_info[sheet_name] = format_info
        all_html_content.append(html_content)

    # 生成完整的HTML内容
    final_html_content = """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        .sheet-container { margin-bottom: 30px; }
        .sheet-title { font-size: 18px; font-weight: bold; margin-bottom: 10px; }
        table { border-collapse: collapse; }
        td { border: 1px solid #000; padding: 5px; }
        .merged-cell { border: 1px solid #000; }
    </style>
</head>
<body>
"""

    # 添加所有工作表的HTML内容
    for html_content in all_html_content:
        final_html_content += html_content + "\n"

    final_html_content += """</body>
</html>"""

    # 保存HTML文件
    with open(html_file, "w", encoding="utf-8") as f:
        f.write(final_html_content)

    # 保存格式信息到元数据文件
    if meta_file:
        import json

        with open(meta_file, "w", encoding="utf-8") as f:
            json.dump(all_format_info, f, indent=2, ensure_ascii=False)

    return all_format_info
